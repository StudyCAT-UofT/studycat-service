#!/usr/bin/env python3
"""
dataset_loader.py

Load a StudyCAT-style tabular dataset into a pandas.DataFrame.
- Supports .csv and .xlsx
- Simple required-column validation
- Exposes a small public API plus a main() for local testing
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, Sequence, Dict, Any, Tuple

import pandas as pd
from openpyxl import load_workbook

# --------------------------- Configuration -----------------------------------
# Update this path locally when you want to test the loader
DATASET_PATH: Path = Path("../../data/utsc-utoronto-ca-2025-09-22_combine.xlsx")


# --------------------------- Exceptions --------------------------------------
class DatasetLoadError(Exception):
    """Raised when the dataset fails to load due to IO or parsing issues."""


class DatasetValidationError(Exception):
    """Raised when the dataset loads but violates expected schema constraints."""


# --------------------------- Helpers -----------------------------------------
DEFAULT_OPTION_LABELS: Tuple[str, ...] = ("A", "B", "C", "D")


def _read_csv(path: Path) -> pd.DataFrame:
    """Read a CSV file into a DataFrame."""
    try:
        return pd.read_csv(path)
    except Exception as e:
        raise DatasetLoadError(f"Failed to read CSV {path!s}: {e!r}")


def _read_excel(path: Path, sheet: Optional[str | int]) -> pd.DataFrame:
    """Read an Excel sheet into a DataFrame. Defaults to the first sheet when sheet is None."""
    effective = 0 if sheet is None else sheet
    try:
        return pd.read_excel(path, sheet_name=effective)
    except Exception as e:
        raise DatasetLoadError(f"Failed to read Excel {path!s}: {e!r}")


def _validate_required_columns(df: pd.DataFrame, required: Optional[Sequence[str]]) -> None:
    if not required:
        return
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DatasetValidationError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Available columns: {list(df.columns)}"
        )


def _normalize_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _extract_bold_correct_indices(
    path: Path,
    *,
    question_column: str,
    option_labels: Sequence[str],
) -> Dict[str, int]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.active

        header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
        header_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_cells):
            title = cell.value
            if isinstance(title, str):
                header_map[title.strip()] = idx

        required_headers = [question_column] + [f"Response_{label}" for label in option_labels]
        missing_headers = [name for name in required_headers if name not in header_map]
        if missing_headers:
            raise DatasetValidationError(
                f"Missing header(s) for bold detection: {missing_headers}"
            )

        lookup: Dict[str, int] = {}
        for row in sheet.iter_rows(min_row=2):
            q_cell = row[header_map[question_column]]
            question_value = q_cell.value
            question_key = _normalize_key(question_value)
            if not question_key:
                continue

            bold_indices: list[int] = []
            for idx, label in enumerate(option_labels):
                cell = row[header_map[f"Response_{label}"]]
                if cell.font and cell.font.bold:
                    bold_indices.append(idx)

            if not bold_indices:
                continue
            if len(bold_indices) > 1:
                raise DatasetValidationError(
                    f"Multiple bold responses found for Question_ID={question_key}: {bold_indices}"
                )

            lookup[question_key] = bold_indices[0]

        return lookup
    finally:
        workbook.close()


# --------------------------- Public API --------------------------------------
def load_dataset(
    path: str | Path | None = None,
    *,
    required_columns: Optional[Sequence[str]] = None,
    excel_sheet: Optional[str] = None,
    derive_correct_from_bold: bool = False,
    question_column: str = "Question_ID",
    option_labels: Sequence[str] = DEFAULT_OPTION_LABELS,
    correct_index_column: str = "__correct_index",
) -> pd.DataFrame:
    """
    Load a dataset file into a pandas DataFrame.

    Parameters
    ----------
    path : str | Path | None
        Path to .csv or .xlsx. If None, uses DATASET_PATH.
    required_columns : Optional[Sequence[str]]
        Column names that must exist in the dataset.
    excel_sheet : Optional[str]
        Sheet name for Excel inputs. Defaults to the first sheet.

    Returns
    -------
    pd.DataFrame
    """
    p = Path(path) if path is not None else DATASET_PATH
    if not p.exists():
        raise DatasetLoadError(f"File not found: {p!s}")
    if not p.is_file():
        raise DatasetLoadError(f"Path is not a file: {p!s}")

    ext = p.suffix.lower()
    if ext == ".csv":
        df = _read_csv(p)
    elif ext == ".xlsx":
        df = _read_excel(p, excel_sheet)
    else:
        raise DatasetLoadError(f"Unsupported file extension {ext!r}. Use .csv or .xlsx.")

    # Normalize column names (strip whitespace)
    df.columns = [str(c).strip() for c in df.columns]

    # Optional required-column validation
    _validate_required_columns(df, required_columns)

    if derive_correct_from_bold:
        if ext != ".xlsx":
            raise DatasetValidationError(
                "derive_correct_from_bold requires an Excel (.xlsx) dataset"
            )
        lookup = _extract_bold_correct_indices(
            p,
            question_column=question_column,
            option_labels=option_labels,
        )

        keys = df[question_column].apply(_normalize_key)
        missing_keys = keys[keys == ""].tolist()
        if missing_keys:
            raise DatasetValidationError(
                f"Row(s) missing {question_column}: {missing_keys}"
            )

        unresolved = [key for key in keys if key not in lookup]
        if unresolved:
            raise DatasetValidationError(
                f"No bold (correct) response found for Question_ID(s): {sorted(set(unresolved))}"
            )

        df[correct_index_column] = keys.map(lookup)

    return df


# --------------------------- Local test entrypoint ----------------------------
def _summarize(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    head_cols = ", ".join(map(str, cols[:8])) + ("..." if len(cols) > 8 else "")
    nn = df[cols[: min(5, len(cols))]].notnull().sum().to_dict() if cols else {}
    pd.set_option("display.max_columns", None)
    return (
        f"Rows: {len(df):,}\n"
        f"Columns: {len(cols):,}\n"
        f"Preview columns: {head_cols}\n"
        f"Non-null (first 5 cols): {nn}\n"
        f"Top 5 rows:\n{df.head()}"
    )


def main() -> None:
    """
    Local smoke-test:
    - Loads DATASET_PATH
    - Prints a short summary
    """
    try:
        df = load_dataset()
    except (DatasetLoadError, DatasetValidationError) as e:
        print(str(e))
        return
    print(_summarize(df))


if __name__ == "__main__":
    main()
