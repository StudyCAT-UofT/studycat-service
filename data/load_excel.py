import pandas as pd
from openpyxl import load_workbook


def excel_to_df(file_path: str) -> pd.DataFrame:
    """
    Converts an excel file in the exact format provided by Prof. De Melo to a pandas dataframe.
    Adds an additional column 'Correct_Answer', specifying which response column has the correct answer.

    Arguments:
        - file_path(str): The file path to the excel file that will be loaded.

    Returns:
        - A pandas DataFrame containing all questions and information, with an additional 'Correct_Answer' column.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Load the data normally (for content)
    df = pd.read_excel(file_path)

    # Find all response columns (e.g., Response_A, Response_B, etc.)
    response_cols = [col for col in df.columns if col.startswith("Response_")]

    # Create a list to store correct answers
    correct_answers = []

    # Iterate over rows in the worksheet (skipping header)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        correct = None
        for col in response_cols:
            # Find corresponding Excel column index
            col_index = list(df.columns).index(col)
            cell = row[col_index]
            if cell.font.bold:  # Check if text is bolded
                correct = col  # Store the column name of the correct answer
                break
        correct_answers.append(correct)

    # Add the new column to the DataFrame
    df["Correct_Answer"] = correct_answers
    return df
